"""Render a GHG inventory spreadsheet as the Markdown the SDK analyzes.

`AekoInventoryAnalyzer.analyze()` takes the inventory "rendered as Markdown — a
table is the natural shape", while the repository reads an `.xlsx` out of S3.
This module is the one place that bridge lives, so the service can stay about
the flow and not about spreadsheets.
"""

from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def inventory_markdown_from_xlsx(data: bytes) -> str:
    """Render every populated sheet of a workbook as a Markdown table.

    The first populated row of a sheet is its header; every later row is a
    data row, padded to the header's width so a short row cannot shift the
    columns under it.

    Raises:
        ValueError: If the bytes are not a readable workbook, or hold no rows.
    """

    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, KeyError, ValueError) as exc:
        raise ValueError(f"The inventory file is not a readable .xlsx: {exc}") from exc

    try:
        sections = [
            section
            for worksheet in workbook.worksheets
            if (section := _sheet_as_markdown(worksheet))
        ]
    finally:
        workbook.close()

    if not sections:
        raise ValueError("The inventory spreadsheet has no data to analyze.")

    return "\n\n".join(sections)


def _sheet_as_markdown(worksheet) -> str:
    rows = [row for row in worksheet.iter_rows(values_only=True) if _has_content(row)]
    if not rows:
        return ""

    header, *body = rows
    width = len(header)

    lines = [
        f"## {worksheet.title}",
        "",
        _as_row(header, width),
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    lines.extend(_as_row(row, width) for row in body)

    return "\n".join(lines)


def _has_content(row) -> bool:
    return any(cell is not None and str(cell).strip() for cell in row)


def _as_row(row, width: int) -> str:
    cells = [_as_cell(row[index] if index < len(row) else None) for index in range(width)]
    return "| " + " | ".join(cells) + " |"


def _as_cell(value) -> str:
    if value is None:
        return ""
    # A pipe would close the column and a line break would end the row, so
    # neither may reach the table as written.
    return " ".join(str(value).split()).replace("|", r"\|")
